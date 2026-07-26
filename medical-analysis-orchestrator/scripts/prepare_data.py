#!/usr/bin/env python3
"""Create a traceable analysis copy without modifying the original dataset."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import unicodedata
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pandas as pd
import yaml
from openpyxl import Workbook
from openpyxl.styles import Border, Font, Side


def utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while chunk := handle.read(1024 * 1024):
            digest.update(chunk)
    return digest.hexdigest()


def load_config(path: Path) -> dict[str, Any]:
    payload = yaml.safe_load(path.read_text(encoding="utf-8-sig"))
    if not isinstance(payload, dict):
        raise ValueError("分析配置根节点必须是映射。")
    return payload


def resolve_path(value: str, config_path: Path) -> Path:
    path = Path(value).expanduser()
    return path.resolve() if path.is_absolute() else (config_path.parent / path).resolve()


def read_frame(path: Path, config: dict[str, Any]) -> pd.DataFrame:
    suffix = path.suffix.lower()
    encoding = (config.get("input") or {}).get("encoding", "auto")
    encoding_value = None if encoding == "auto" else encoding
    sheet = (config.get("input") or {}).get("sheet")
    if suffix == ".csv":
        return pd.read_csv(path, encoding=encoding_value, low_memory=False)
    if suffix in {".tsv", ".txt", ".dat"}:
        separator = "\t" if suffix == ".tsv" else None
        return pd.read_csv(
            path,
            sep=separator,
            engine="python",
            encoding=encoding_value,
        )
    if suffix in {".xlsx", ".xls"}:
        return pd.read_excel(path, sheet_name=sheet if sheet is not None else 0)
    if suffix == ".sav":
        return pd.read_spss(path)
    if suffix == ".dta":
        return pd.read_stata(path)
    if suffix in {".sas7bdat", ".xpt"}:
        return pd.read_sas(path, format="xport" if suffix == ".xpt" else None)
    if suffix == ".json":
        return pd.read_json(path)
    if suffix == ".jsonl":
        return pd.read_json(path, lines=True)
    if suffix == ".parquet":
        return pd.read_parquet(path)
    if suffix == ".feather":
        return pd.read_feather(path)
    raise ValueError(f"暂不支持正式准备此格式：{suffix}")


def normalized_name(raw: Any, position: int) -> str:
    value = unicodedata.normalize("NFKC", str(raw)).strip()
    value = re.sub(r"[\u0000-\u001f\u007f-\u009f]", "", value)
    value = re.sub(r"\s+", "_", value)
    value = re.sub(r"[^\w]+", "_", value, flags=re.UNICODE).strip("_")
    return value or f"unnamed_{position}"


def unique_names(raw_names: list[Any]) -> tuple[list[str], list[dict[str, Any]]]:
    result: list[str] = []
    mapping: list[dict[str, Any]] = []
    counts: dict[str, int] = {}
    for position, raw in enumerate(raw_names, start=1):
        base = normalized_name(raw, position)
        counts[base] = counts.get(base, 0) + 1
        clean = base if counts[base] == 1 else f"{base}__{counts[base]}"
        result.append(clean)
        mapping.append(
            {
                "column_position": position,
                "original_name": str(raw),
                "analysis_name": clean,
                "changed": clean != str(raw),
            }
        )
    return result, mapping


def compare_mask(series: pd.Series, operator: str, value: Any) -> pd.Series:
    if operator == "eq":
        return series == value
    if operator == "ne":
        return series != value
    if operator == "lt":
        return pd.to_numeric(series, errors="coerce") < float(value)
    if operator == "le":
        return pd.to_numeric(series, errors="coerce") <= float(value)
    if operator == "gt":
        return pd.to_numeric(series, errors="coerce") > float(value)
    if operator == "ge":
        return pd.to_numeric(series, errors="coerce") >= float(value)
    if operator == "is_missing":
        return series.isna()
    raise ValueError(f"不支持的排除运算符：{operator}")


def apply_confirmed_action(
    frame: pd.DataFrame,
    action: dict[str, Any],
) -> tuple[pd.DataFrame, int]:
    action_type = str(action.get("action", ""))
    if not action.get("confirmed_by") or not action.get("confirmed_at"):
        raise ValueError(f"清洗操作 {action.get('id')} 缺少确认人或确认时间。")
    before = len(frame)
    if action_type == "recode_to_missing":
        variables = [str(item) for item in action.get("variables", [])]
        values = action.get("values", [])
        affected = 0
        for variable in variables:
            if variable not in frame:
                raise ValueError(f"清洗变量不存在：{variable}")
            mask = frame[variable].isin(values)
            affected += int(mask.sum())
            frame.loc[mask, variable] = pd.NA
    elif action_type == "recode_values":
        variable = str(action.get("variable", ""))
        mapping = action.get("mapping") or {}
        if variable not in frame:
            raise ValueError(f"清洗变量不存在：{variable}")
        mask = frame[variable].isin(mapping.keys())
        affected = int(mask.sum())
        frame.loc[mask, variable] = frame.loc[mask, variable].map(mapping)
    elif action_type == "drop_duplicates":
        subset = action.get("variables") or None
        frame = frame.drop_duplicates(subset=subset, keep=str(action.get("keep", "first")))
        affected = before - len(frame)
    elif action_type == "exclude_rows":
        variable = str(action.get("variable", ""))
        if variable not in frame:
            raise ValueError(f"清洗变量不存在：{variable}")
        mask = compare_mask(frame[variable], str(action.get("operator", "eq")), action.get("value"))
        affected = int(mask.sum())
        frame = frame.loc[~mask].copy()
    elif action_type == "convert_type":
        variable = str(action.get("variable", ""))
        target = str(action.get("target_type", ""))
        if variable not in frame:
            raise ValueError(f"清洗变量不存在：{variable}")
        before_values = frame[variable].copy()
        if target == "numeric":
            frame[variable] = pd.to_numeric(frame[variable], errors="coerce")
        elif target == "string":
            frame[variable] = frame[variable].astype("string")
        elif target == "category":
            frame[variable] = frame[variable].astype("category")
        elif target == "datetime":
            frame[variable] = pd.to_datetime(frame[variable], errors="coerce")
        else:
            raise ValueError(f"不支持的目标类型：{target}")
        before_text = before_values.astype("string").fillna("<NA>")
        after_text = frame[variable].astype("string").fillna("<NA>")
        affected = int((before_text != after_text).sum())
    else:
        raise ValueError(f"清洗操作尚未实现，拒绝执行：{action_type}")
    return frame, affected


def write_clean_xlsx(frame: pd.DataFrame, path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "清洁分析数据"
    worksheet.sheet_view.showGridLines = False
    worksheet.append(list(frame.columns))
    for row in frame.itertuples(index=False, name=None):
        worksheet.append([None if pd.isna(value) else value for value in row])
    top = Side(style="medium", color="000000")
    thin = Side(style="thin", color="000000")
    bottom = Side(style="medium", color="000000")
    for cell in worksheet[1]:
        cell.font = Font(name="宋体", size=10.5, bold=True)
        cell.border = Border(top=top, bottom=thin)
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(
                name="Times New Roman"
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool)
                else "宋体",
                size=10.5,
            )
    if worksheet.max_row >= 2:
        for cell in worksheet[worksheet.max_row]:
            cell.border = Border(bottom=bottom)
    worksheet.freeze_panes = "A2"
    for cells in worksheet.columns:
        width = min(
            35,
            max(10, max(len(str(cell.value)) if cell.value is not None else 0 for cell in cells) + 2),
        )
        worksheet.column_dimensions[cells[0].column_letter].width = width
    workbook.save(path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="生成只读原始数据对应的清洁分析副本。")
    parser.add_argument("--config", required=True)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    config_path = Path(args.config).expanduser().resolve()
    config = load_config(config_path)
    approval = config.get("approval") or {}
    if approval.get("confirmed") is not True:
        raise SystemExit("数据准备只能在分析方案确认后执行。")
    input_path = resolve_path(str((config.get("input") or {}).get("path", "")), config_path)
    if not input_path.is_file():
        raise SystemExit("正式执行时 input.path 必须指向单个文件。")
    expected = (config.get("input") or {}).get("expected_sha256")
    before_hash = sha256_file(input_path)
    if expected and str(expected).lower() != before_hash.lower():
        raise SystemExit("原始数据指纹与确认方案不一致。")

    run_dir_value = str((config.get("run") or {}).get("output_dir") or config_path.parent)
    run_dir = resolve_path(run_dir_value, config_path)
    data_dir = run_dir / "01_数据整理"
    data_dir.mkdir(parents=True, exist_ok=True)

    frame = read_frame(input_path, config)
    input_rows, input_columns = frame.shape
    clean_names, name_mapping = unique_names(list(frame.columns))
    frame.columns = clean_names
    log_rows: list[dict[str, Any]] = []
    changed_names = sum(item["changed"] for item in name_mapping)
    log_rows.append(
        {
            "operation_id": "auto_normalize_column_names",
            "action": "normalize_column_names",
            "scope": "all_columns",
            "affected_count": changed_names,
            "confirmation_required": False,
            "confirmed_by": "automatic_policy",
            "status": "completed",
            "executed_at_utc": utc_now(),
            "reason": "只修改清洁副本并保留原始名称映射。",
        }
    )

    actions = (config.get("data_handling") or {}).get("confirmed_actions") or []
    for action in actions:
        frame, affected = apply_confirmed_action(frame, action)
        log_rows.append(
            {
                "operation_id": action.get("id"),
                "action": action.get("action"),
                "scope": ",".join(map(str, action.get("variables", [action.get("variable", "")]))),
                "affected_count": affected,
                "confirmation_required": True,
                "confirmed_by": action.get("confirmed_by"),
                "status": "completed",
                "executed_at_utc": utc_now(),
                "reason": action.get("reason", ""),
            }
        )

    csv_path = data_dir / "05_清洁分析数据.csv"
    xlsx_path = data_dir / "06_清洁分析数据.xlsx"
    log_path = data_dir / "04_数据清洗日志.csv"
    mapping_path = data_dir / "04_变量名称映射.csv"
    frame.to_csv(csv_path, index=False, encoding="utf-8-sig")
    write_clean_xlsx(frame, xlsx_path)
    pd.DataFrame(name_mapping).to_csv(mapping_path, index=False, encoding="utf-8-sig")
    with log_path.open("w", encoding="utf-8-sig", newline="") as handle:
        fields = [
            "operation_id",
            "action",
            "scope",
            "affected_count",
            "confirmation_required",
            "confirmed_by",
            "status",
            "executed_at_utc",
            "reason",
        ]
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(log_rows)

    summary = {
        "schema_version": "1.0",
        "created_at_utc": utc_now(),
        "source_path": str(input_path),
        "source_sha256": before_hash,
        "input_rows": input_rows,
        "input_columns": input_columns,
        "output_rows": len(frame),
        "output_columns": len(frame.columns),
        "confirmed_action_count": len(actions),
        "clean_csv": str(csv_path),
        "clean_csv_sha256": sha256_file(csv_path),
        "clean_xlsx": str(xlsx_path),
        "clean_xlsx_sha256": sha256_file(xlsx_path),
    }
    (data_dir / "data_preparation_summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(json.dumps(summary, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
