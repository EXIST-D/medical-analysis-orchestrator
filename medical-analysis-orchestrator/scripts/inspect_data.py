#!/usr/bin/env python3
"""Read-only medical dataset inventory and aggregate profiling."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import re
import statistics
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

try:
    import pandas as pd
except ImportError:  # CSV/JSON still work through the standard library.
    pd = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Border, Font, Side
except ImportError:
    Workbook = None
    Border = Font = Side = None


SUPPORTED_EXTENSIONS = {
    ".csv",
    ".tsv",
    ".txt",
    ".dat",
    ".json",
    ".jsonl",
    ".xlsx",
    ".xls",
    ".sav",
    ".dta",
    ".sas7bdat",
    ".xpt",
    ".parquet",
    ".feather",
}
INVENTORY_ONLY_EXTENSIONS = {".rds", ".rda", ".rdata"}

ROLE_PATTERNS = {
    "id": re.compile(
        r"(^|_)(id|patient_id|subject_id|record_id|case_id|mrn|病历号|编号)(_|$)",
        re.IGNORECASE,
    ),
    "outcome": re.compile(
        r"(outcome|endpoint|response|status|result|结局|终点|疗效|转归)",
        re.IGNORECASE,
    ),
    "group": re.compile(
        r"(^|_)(group|arm|treatment|cohort|class|组别|分组|队列)(_|$)",
        re.IGNORECASE,
    ),
    "time": re.compile(
        r"(date|time|day|month|year|visit|follow.?up|日期|时间|随访|访视)",
        re.IGNORECASE,
    ),
    "event": re.compile(
        r"(^|_)(event|death|relapse|failure|censor|死亡|复发|事件)(_|$)",
        re.IGNORECASE,
    ),
    "scale_item": re.compile(
        r"(^|_)(item|question|q|phq|gad|sf|score_item)[_-]?\d+($|_)|"
        r"(条目|题目)\d+",
        re.IGNORECASE,
    ),
}

DIRECT_IDENTIFIER_PATTERN = re.compile(
    r"(name|姓名|身份证|idcard|phone|mobile|电话|email|邮箱|address|地址|mrn|病历号)",
    re.IGNORECASE,
)


def utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def sha256_file(path: Path, chunk_size: int = 1024 * 1024) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while chunk := handle.read(chunk_size):
            digest.update(chunk)
    return digest.hexdigest()


def discover_files(input_path: Path, recursive: bool) -> list[Path]:
    if input_path.is_file():
        return [input_path]
    iterator: Iterable[Path]
    iterator = input_path.rglob("*") if recursive else input_path.glob("*")
    return sorted(path for path in iterator if path.is_file())


def _read_with_pandas(path: Path, encoding: str) -> Any:
    suffix = path.suffix.lower()
    csv_encoding = None if encoding == "auto" else encoding
    if suffix == ".csv":
        return pd.read_csv(path, encoding=csv_encoding, low_memory=False)
    if suffix in {".tsv", ".txt", ".dat"}:
        separator = "\t" if suffix == ".tsv" else None
        return pd.read_csv(
            path,
            sep=separator,
            engine="python" if separator is None else "c",
            encoding=csv_encoding,
        )
    if suffix == ".json":
        return pd.read_json(path)
    if suffix == ".jsonl":
        return pd.read_json(path, lines=True)
    if suffix in {".xlsx", ".xls"}:
        return pd.read_excel(path)
    if suffix == ".sav":
        return pd.read_spss(path)
    if suffix == ".dta":
        return pd.read_stata(path)
    if suffix in {".sas7bdat", ".xpt"}:
        return pd.read_sas(path, format="xport" if suffix == ".xpt" else None)
    if suffix == ".parquet":
        return pd.read_parquet(path)
    if suffix == ".feather":
        return pd.read_feather(path)
    raise ValueError(f"Unsupported readable format: {suffix}")


def _read_with_stdlib(path: Path, encoding: str) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    chosen_encoding = "utf-8-sig" if encoding == "auto" else encoding
    if suffix in {".csv", ".tsv", ".txt", ".dat"}:
        with path.open("r", encoding=chosen_encoding, newline="") as handle:
            if suffix == ".csv":
                dialect = csv.excel
            elif suffix == ".tsv":
                dialect = csv.excel_tab
            else:
                sample = handle.read(8192)
                handle.seek(0)
                dialect = csv.Sniffer().sniff(sample)
            return list(csv.DictReader(handle, dialect=dialect))
    if suffix in {".json", ".jsonl"}:
        with path.open("r", encoding=chosen_encoding) as handle:
            if suffix == ".jsonl":
                return [json.loads(line) for line in handle if line.strip()]
            payload = json.load(handle)
        if isinstance(payload, list) and all(isinstance(item, dict) for item in payload):
            return payload
        raise ValueError("Standard-library JSON reader requires a list of objects")
    raise RuntimeError(
        f"Reading {suffix} requires pandas and the relevant optional engine"
    )


def read_dataset(path: Path, encoding: str) -> Any:
    return (
        _read_with_pandas(path, encoding)
        if pd is not None
        else _read_with_stdlib(path, encoding)
    )


def read_datasets(path: Path, encoding: str) -> list[tuple[str, Any]]:
    """Return one or more named tabular datasets from a file."""
    if pd is not None and path.suffix.lower() in {".xlsx", ".xls"}:
        sheets = pd.read_excel(path, sheet_name=None)
        return [(f"{path.name}::{sheet}", frame) for sheet, frame in sheets.items()]
    return [(path.name, read_dataset(path, encoding))]


def _safe_number(value: Any) -> float | int | None:
    try:
        if value is None or (isinstance(value, float) and math.isnan(value)):
            return None
        numeric = float(value)
        return int(numeric) if numeric.is_integer() else numeric
    except (TypeError, ValueError, OverflowError):
        return None


def _profile_with_pandas(frame: Any, source: str) -> dict[str, Any]:
    rows, columns = frame.shape
    duplicate_rows = int(frame.duplicated().sum())
    variable_profiles: list[dict[str, Any]] = []
    name_counts: dict[str, int] = {}
    for raw_name in frame.columns:
        text_name = str(raw_name)
        name_counts[text_name] = name_counts.get(text_name, 0) + 1

    for position, raw_name in enumerate(frame.columns):
        name = str(raw_name)
        series = frame.iloc[:, position]
        non_missing = series.dropna()
        missing_n = int(series.isna().sum())
        unique_n = int(non_missing.nunique(dropna=True))
        unique_ratio = unique_n / len(non_missing) if len(non_missing) else 0.0
        roles = [role for role, pattern in ROLE_PATTERNS.items() if pattern.search(name)]

        if pd.api.types.is_datetime64_any_dtype(series):
            inferred_type = "datetime"
            if "time" not in roles:
                roles.append("time")
        elif pd.api.types.is_bool_dtype(series) or unique_n == 2:
            inferred_type = "binary"
        elif pd.api.types.is_numeric_dtype(series):
            category_limit = min(20, max(3, int(math.sqrt(max(rows, 1)))))
            inferred_type = "categorical_numeric" if unique_n <= category_limit else "continuous"
        elif isinstance(series.dtype, pd.CategoricalDtype):
            inferred_type = "categorical"
        elif unique_n <= min(30, max(3, int(math.sqrt(max(rows, 1))))):
            inferred_type = "categorical"
        else:
            inferred_type = "text"

        if (
            "id" not in roles
            and rows > 0
            and unique_ratio >= 0.98
            and inferred_type == "text"
        ):
            roles.append("possible_id")

        numeric_summary = None
        suppress_summary = (
            "id" in roles
            or "possible_id" in roles
            or bool(DIRECT_IDENTIFIER_PATTERN.search(name))
        )
        if (
            pd.api.types.is_numeric_dtype(series)
            and len(non_missing)
            and not suppress_summary
        ):
            numeric = pd.to_numeric(non_missing, errors="coerce").dropna()
            if len(numeric):
                quantiles = numeric.quantile([0.25, 0.5, 0.75])
                numeric_summary = {
                    "min": _safe_number(numeric.min()),
                    "q1": _safe_number(quantiles.loc[0.25]),
                    "median": _safe_number(quantiles.loc[0.5]),
                    "q3": _safe_number(quantiles.loc[0.75]),
                    "max": _safe_number(numeric.max()),
                    "mean": _safe_number(numeric.mean()),
                    "sd": _safe_number(numeric.std()) if len(numeric) > 1 else None,
                }

        variable_profiles.append(
            {
                "name": name,
                "column_position": position + 1,
                "duplicate_column_name": name_counts[name] > 1,
                "source_dtype": str(series.dtype),
                "inferred_type": inferred_type,
                "missing_n": missing_n,
                "missing_pct": round(missing_n / rows * 100, 4) if rows else 0.0,
                "non_missing_n": int(len(non_missing)),
                "unique_n": unique_n,
                "unique_ratio": round(unique_ratio, 6),
                "constant": unique_n <= 1,
                "candidate_roles": sorted(set(roles)),
                "potential_direct_identifier": bool(
                    DIRECT_IDENTIFIER_PATTERN.search(name)
                ),
                "numeric_summary": numeric_summary,
            }
        )

    return {
        "source": source,
        "rows": int(rows),
        "columns": int(columns),
        "duplicate_rows": duplicate_rows,
        "duplicate_pct": round(duplicate_rows / rows * 100, 4) if rows else 0.0,
        "variables": variable_profiles,
    }


def _profile_with_stdlib(records: list[dict[str, Any]], source: str) -> dict[str, Any]:
    field_names = sorted({key for row in records for key in row})
    normalized_rows = [
        tuple((key, json.dumps(row.get(key), ensure_ascii=False, sort_keys=True)) for key in field_names)
        for row in records
    ]
    duplicate_rows = len(normalized_rows) - len(set(normalized_rows))
    variables = []
    for name in field_names:
        values = [row.get(name) for row in records]
        non_missing = [value for value in values if value not in (None, "")]
        unique_n = len({json.dumps(value, ensure_ascii=False, sort_keys=True) for value in non_missing})
        roles = [role for role, pattern in ROLE_PATTERNS.items() if pattern.search(name)]
        numeric_values: list[float] = []
        all_numeric = bool(non_missing)
        for value in non_missing:
            try:
                numeric_values.append(float(value))
            except (TypeError, ValueError):
                all_numeric = False
                numeric_values = []
                break
        category_limit = min(20, max(3, int(math.sqrt(max(len(records), 1)))))
        if unique_n == 2:
            inferred_type = "binary"
        elif all_numeric:
            inferred_type = (
                "categorical_numeric"
                if unique_n <= category_limit
                else "continuous"
            )
        elif unique_n <= min(30, max(3, int(math.sqrt(max(len(records), 1))))):
            inferred_type = "categorical"
        else:
            inferred_type = "text"
        unique_ratio = unique_n / len(non_missing) if non_missing else 0.0
        if (
            "id" not in roles
            and unique_ratio >= 0.98
            and inferred_type == "text"
        ):
            roles.append("possible_id")
        numeric_summary = None
        suppress_summary = (
            "id" in roles
            or "possible_id" in roles
            or bool(DIRECT_IDENTIFIER_PATTERN.search(name))
        )
        if numeric_values and not suppress_summary:
            ordered = sorted(numeric_values)

            def quantile(probability: float) -> float:
                if len(ordered) == 1:
                    return ordered[0]
                position = (len(ordered) - 1) * probability
                lower = math.floor(position)
                upper = math.ceil(position)
                if lower == upper:
                    return ordered[lower]
                weight = position - lower
                return ordered[lower] * (1 - weight) + ordered[upper] * weight

            numeric_summary = {
                "min": _safe_number(ordered[0]),
                "q1": _safe_number(quantile(0.25)),
                "median": _safe_number(quantile(0.5)),
                "q3": _safe_number(quantile(0.75)),
                "max": _safe_number(ordered[-1]),
                "mean": _safe_number(statistics.fmean(ordered)),
                "sd": _safe_number(statistics.stdev(ordered))
                if len(ordered) > 1
                else None,
            }
        variables.append(
            {
                "name": name,
                "column_position": field_names.index(name) + 1,
                "duplicate_column_name": False,
                "source_dtype": "numeric_text" if all_numeric else "text",
                "inferred_type": inferred_type,
                "missing_n": len(values) - len(non_missing),
                "missing_pct": round(
                    (len(values) - len(non_missing)) / len(values) * 100, 4
                )
                if values
                else 0.0,
                "non_missing_n": len(non_missing),
                "unique_n": unique_n,
                "unique_ratio": round(unique_ratio, 6),
                "constant": unique_n <= 1,
                "candidate_roles": sorted(set(roles)),
                "potential_direct_identifier": bool(
                    DIRECT_IDENTIFIER_PATTERN.search(name)
                ),
                "numeric_summary": numeric_summary,
            }
        )
    return {
        "source": source,
        "rows": len(records),
        "columns": len(field_names),
        "duplicate_rows": duplicate_rows,
        "duplicate_pct": round(duplicate_rows / len(records) * 100, 4)
        if records
        else 0.0,
        "variables": variables,
        "warnings": ["pandas_not_available_type_inference_limited"],
    }


def profile_dataset(data: Any, source: str) -> dict[str, Any]:
    if pd is not None:
        return _profile_with_pandas(data, source)
    return _profile_with_stdlib(data, source)


def flatten_profiles(profiles: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    for dataset in profiles:
        for variable in dataset["variables"]:
            summary = variable.get("numeric_summary") or {}
            rows.append(
                {
                    "source": dataset["source"],
                    "dataset_rows": dataset["rows"],
                    "dataset_columns": dataset["columns"],
                    "variable": variable["name"],
                    "column_position": variable["column_position"],
                    "duplicate_column_name": variable["duplicate_column_name"],
                    "source_dtype": variable["source_dtype"],
                    "inferred_type": variable["inferred_type"],
                    "missing_n": variable["missing_n"],
                    "missing_pct": variable["missing_pct"],
                    "unique_n": variable["unique_n"],
                    "unique_ratio": variable["unique_ratio"],
                    "constant": variable["constant"],
                    "candidate_roles": "|".join(variable["candidate_roles"]),
                    "potential_direct_identifier": variable[
                        "potential_direct_identifier"
                    ],
                    "min": summary.get("min"),
                    "q1": summary.get("q1"),
                    "median": summary.get("median"),
                    "q3": summary.get("q3"),
                    "max": summary.get("max"),
                    "mean": summary.get("mean"),
                    "sd": summary.get("sd"),
                }
            )
    return rows


def write_json(path: Path, payload: Any) -> None:
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def write_profile_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    field_names = [
        "source",
        "dataset_rows",
        "dataset_columns",
        "variable",
        "column_position",
        "duplicate_column_name",
        "source_dtype",
        "inferred_type",
        "missing_n",
        "missing_pct",
        "unique_n",
        "unique_ratio",
        "constant",
        "candidate_roles",
        "potential_direct_identifier",
        "min",
        "q1",
        "median",
        "q3",
        "max",
        "mean",
        "sd",
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=field_names)
        writer.writeheader()
        writer.writerows(rows)


def artifact_record(path: Path, output_dir: Path, artifact_type: str) -> dict[str, Any]:
    return {
        "path": path.relative_to(output_dir).as_posix(),
        "type": artifact_type,
        "bytes": path.stat().st_size,
        "sha256": sha256_file(path),
        "created_at_utc": utc_now(),
        "module": "inspect",
    }


def _excel_safe(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple, set)):
        return json.dumps(value, ensure_ascii=False)
    return value


def write_three_line_workbook(
    path: Path,
    sheets: list[tuple[str, list[dict[str, Any]]]],
) -> None:
    if Workbook is None:
        raise RuntimeError("openpyxl is required to create inspection workbooks")
    workbook = Workbook()
    workbook.remove(workbook.active)
    top = Side(style="medium", color="000000")
    thin = Side(style="thin", color="000000")
    bottom = Side(style="medium", color="000000")
    for sheet_name, rows in sheets:
        worksheet = workbook.create_sheet(title=sheet_name[:31] or "结果")
        worksheet.sheet_view.showGridLines = False
        field_names = list(rows[0].keys()) if rows else ["message"]
        worksheet.append(field_names)
        for row in rows:
            worksheet.append([_excel_safe(row.get(field)) for field in field_names])
        for cell in worksheet[1]:
            cell.font = Font(name="宋体", size=10.5, bold=True)
            cell.border = Border(top=top, bottom=thin)
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    cell.font = Font(name="Times New Roman", size=10.5)
                else:
                    cell.font = Font(name="宋体", size=10.5)
        if worksheet.max_row >= 2:
            for cell in worksheet[worksheet.max_row]:
                cell.border = Border(bottom=bottom)
        worksheet.freeze_panes = "A2"
        for column_cells in worksheet.columns:
            width = max(
                10,
                min(
                    40,
                    max(
                        len(str(cell.value)) if cell.value is not None else 0
                        for cell in column_cells
                    )
                    + 2,
                ),
            )
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def cleaning_candidates(profiles: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for dataset in profiles:
        source = dataset["source"]
        if dataset.get("duplicate_rows", 0):
            rows.append(
                {
                    "dataset": source,
                    "issue": "duplicate_rows",
                    "variable": "",
                    "count": dataset["duplicate_rows"],
                    "default_action": "report_only",
                    "confirmation_required": True,
                    "reason": "重复记录可能是错误，也可能是合法重复测量。",
                }
            )
        for variable in dataset["variables"]:
            if variable.get("duplicate_column_name"):
                rows.append(
                    {
                        "dataset": source,
                        "issue": "duplicate_column_name",
                        "variable": variable["name"],
                        "count": 1,
                        "default_action": "disambiguate_name_in_clean_copy",
                        "confirmation_required": False,
                        "reason": "只修改清洁副本中的分析变量名并保留映射。",
                    }
                )
            if variable.get("constant"):
                rows.append(
                    {
                        "dataset": source,
                        "issue": "constant_column",
                        "variable": variable["name"],
                        "count": 1,
                        "default_action": "report_only",
                        "confirmation_required": True,
                        "reason": "是否排除常量列由分析方案决定。",
                    }
                )
            if variable.get("potential_direct_identifier"):
                rows.append(
                    {
                        "dataset": source,
                        "issue": "potential_identifier",
                        "variable": variable["name"],
                        "count": 1,
                        "default_action": "report_only",
                        "confirmation_required": True,
                        "reason": "正式报告前需确认去标识化和输出范围。",
                    }
                )
            if variable.get("missing_n", 0):
                rows.append(
                    {
                        "dataset": source,
                        "issue": "missing_values",
                        "variable": variable["name"],
                        "count": variable["missing_n"],
                        "default_action": "report_only",
                        "confirmation_required": True,
                        "reason": "缺失值策略必须按研究问题和模型确认。",
                    }
                )
    return rows


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Read-only inventory and aggregate profiling for medical datasets."
    )
    parser.add_argument("input", help="Data file or directory")
    parser.add_argument("--output", required=True, help="Run output directory")
    parser.add_argument("--recursive", action="store_true", help="Recurse into subdirectories")
    parser.add_argument("--encoding", default="auto", help="Text encoding or auto")
    parser.add_argument(
        "--max-file-mb",
        type=float,
        default=1024.0,
        help="Skip loading files larger than this limit; inventory still records them",
    )
    parser.add_argument(
        "--no-hash",
        action="store_true",
        help="Skip source SHA-256 hashing (not recommended for formal execution)",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    input_path = Path(args.input).expanduser().resolve()
    output_dir = Path(args.output).expanduser().resolve()
    if not input_path.exists():
        print(f"Input path does not exist: {input_path}", file=sys.stderr)
        return 2

    output_dir.mkdir(parents=True, exist_ok=True)
    all_files = discover_files(input_path, args.recursive)
    inventory = []
    profiles = []
    warnings = []

    for path in all_files:
        suffix = path.suffix.lower()
        relative_source = (
            path.name
            if input_path.is_file()
            else path.relative_to(input_path).as_posix()
        )
        size_bytes = path.stat().st_size
        record = {
            "source": relative_source,
            "extension": suffix,
            "size_bytes": size_bytes,
            "modified_at_utc": datetime.fromtimestamp(
                path.stat().st_mtime, timezone.utc
            )
            .replace(microsecond=0)
            .isoformat(),
            "sha256": None if args.no_hash else sha256_file(path),
            "supported": suffix in SUPPORTED_EXTENSIONS,
            "inventory_only": suffix in INVENTORY_ONLY_EXTENSIONS,
            "read_status": "not_attempted",
            "error": None,
        }
        if suffix in INVENTORY_ONLY_EXTENSIONS:
            record["read_status"] = "requires_r_reader"
        elif suffix not in SUPPORTED_EXTENSIONS:
            record["read_status"] = "unsupported"
        elif size_bytes > args.max_file_mb * 1024 * 1024:
            record["read_status"] = "skipped_size_limit"
            record["error"] = f"File exceeds --max-file-mb={args.max_file_mb}"
            warnings.append(f"{relative_source}: skipped because of size limit")
        else:
            try:
                loaded_datasets = read_datasets(path, args.encoding)
                for dataset_name, data in loaded_datasets:
                    source_name = (
                        dataset_name
                        if path.suffix.lower() in {".xlsx", ".xls"}
                        else relative_source
                    )
                    profiles.append(profile_dataset(data, source_name))
                record["read_status"] = "profiled"
                record["datasets"] = [item[0] for item in loaded_datasets]
            except Exception as exc:  # Keep inventory even when an optional reader fails.
                record["read_status"] = "failed"
                record["error"] = f"{type(exc).__name__}: {exc}"
                warnings.append(f"{relative_source}: {record['error']}")
        inventory.append(record)

    role_counts: dict[str, int] = {}
    total_variables = 0
    total_missing_cells = 0
    total_cells = 0
    for dataset in profiles:
        total_variables += dataset["columns"]
        for variable in dataset["variables"]:
            total_missing_cells += variable["missing_n"]
            total_cells += dataset["rows"]
            for role in variable["candidate_roles"]:
                role_counts[role] = role_counts.get(role, 0) + 1

    profile_payload = {
        "schema_version": "1.0",
        "generated_at_utc": utc_now(),
        "input_path": str(input_path),
        "privacy": {
            "contains_patient_level_values": False,
            "direct_identifier_values_exported": False,
        },
        "summary": {
            "file_count": len(all_files),
            "profiled_dataset_count": len(profiles),
            "total_rows_across_datasets": sum(item["rows"] for item in profiles),
            "total_variables_across_datasets": total_variables,
            "overall_missing_pct": round(total_missing_cells / total_cells * 100, 4)
            if total_cells
            else None,
            "candidate_role_counts": role_counts,
        },
        "datasets": profiles,
        "warnings": warnings,
    }
    inventory_payload = {
        "schema_version": "1.0",
        "generated_at_utc": utc_now(),
        "input_path": str(input_path),
        "files": inventory,
    }

    data_dir = output_dir / "01_数据整理"
    data_dir.mkdir(parents=True, exist_ok=True)
    inventory_path = output_dir / "data_inventory.json"
    profile_json_path = output_dir / "data_profile.json"
    profile_csv_path = output_dir / "data_profile.csv"
    dictionary_path = data_dir / "01_变量字典.xlsx"
    quality_path = data_dir / "02_数据质量报告.xlsx"
    cleaning_candidates_path = data_dir / "03_清洗操作候选.csv"
    write_json(inventory_path, inventory_payload)
    write_json(profile_json_path, profile_payload)
    flattened = flatten_profiles(profiles)
    write_profile_csv(profile_csv_path, flattened)
    candidate_rows = cleaning_candidates(profiles)
    with cleaning_candidates_path.open("w", encoding="utf-8-sig", newline="") as handle:
        fields = [
            "dataset",
            "issue",
            "variable",
            "count",
            "default_action",
            "confirmation_required",
            "reason",
        ]
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(candidate_rows)
    try:
        dictionary_rows = [
            {
                "dataset": row["source"],
                "column_position": row["column_position"],
                "original_name": row["variable"],
                "analysis_name": row["variable"],
                "label": "",
                "value_labels": "",
                "metadata_status": "unavailable_in_tabular_source",
                "source_dtype": row["source_dtype"],
                "inferred_type": row["inferred_type"],
                "candidate_roles": row["candidate_roles"],
                "unit": "",
                "missing_n": row["missing_n"],
                "missing_pct": row["missing_pct"],
                "unique_n": row["unique_n"],
                "potential_identifier": row["potential_direct_identifier"],
                "constant": row["constant"],
                "needs_confirmation": bool(
                    row["candidate_roles"]
                    or row["potential_direct_identifier"]
                    or row["duplicate_column_name"]
                ),
            }
            for row in flattened
        ]
        quality_summary = [
            {
                "dataset": item["source"],
                "rows": item["rows"],
                "columns": item["columns"],
                "duplicate_rows": item["duplicate_rows"],
                "duplicate_pct": item["duplicate_pct"],
            }
            for item in profiles
        ]
        write_three_line_workbook(
            dictionary_path,
            [("变量字典", dictionary_rows)],
        )
        write_three_line_workbook(
            quality_path,
            [
                ("数据集概况", quality_summary),
                ("质量问题", candidate_rows),
                ("文件清单", inventory),
            ],
        )
    except Exception as exc:
        warnings.append(f"inspection_workbook_failed: {type(exc).__name__}: {exc}")

    manifest = {
        "schema_version": "1.0",
        "run_id": output_dir.name,
        "phase": "inspect",
        "created_at_utc": utc_now(),
        "input_fingerprints": [
            {
                "source": item["source"],
                "sha256": item["sha256"],
                "size_bytes": item["size_bytes"],
            }
            for item in inventory
        ],
        "analysis_plan_sha256": None,
        "random_seed": None,
        "artifacts": [
            artifact_record(inventory_path, output_dir, "inventory"),
            artifact_record(profile_json_path, output_dir, "profile"),
            artifact_record(profile_csv_path, output_dir, "profile_table"),
            artifact_record(cleaning_candidates_path, output_dir, "cleaning_candidates"),
        ],
        "warnings": warnings,
        "validation": {"status": "pending", "checked_at_utc": None},
    }
    for optional_path, artifact_type in (
        (dictionary_path, "variable_dictionary"),
        (quality_path, "quality_report"),
    ):
        if optional_path.exists():
            manifest["artifacts"].append(
                artifact_record(optional_path, output_dir, artifact_type)
            )
    write_json(output_dir / "manifest.json", manifest)

    print(
        json.dumps(
            {
                "file_count": len(all_files),
                "profiled_dataset_count": len(profiles),
                "output_dir": str(output_dir),
                "warnings": len(warnings),
            },
            ensure_ascii=False,
        )
    )
    return 0 if profiles else 3


if __name__ == "__main__":
    raise SystemExit(main())
